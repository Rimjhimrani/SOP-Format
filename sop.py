import streamlit as st
import io
import datetime
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas as rl_canvas
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Line Supply SOP", page_icon="🏭", layout="wide")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
.main-header {
    background: linear-gradient(135deg, #003366 0%, #0055aa 100%);
    color: white; padding: 18px 28px; border-radius: 10px; margin-bottom: 22px;
}
.main-header h2 { margin:0; font-size:20px; font-weight:700; }
.main-header p  { margin:4px 0 0; font-size:12px; opacity:.8; }
.section-box {
    background: white; border: 1px solid #dde3ec;
    border-top: 4px solid #0055aa; border-radius: 8px;
    padding: 18px 20px; margin-bottom: 18px;
    box-shadow: 0 2px 6px rgba(0,0,0,.06);
}
.section-title { font-size:14px; font-weight:700; color:#003366; margin-bottom:12px; }
.flow-step {
    background:#f0f6ff; border-left:4px solid #0055aa;
    border-radius:5px; padding:10px 14px; margin:6px 0; font-size:13px;
}
.flow-decision {
    background:#fdf5ff; border-left:4px solid #7c3aed;
    border-radius:5px; padding:10px 14px; margin:6px 0; font-size:13px;
}
.flow-arrow { text-align:center; color:#aaa; font-size:16px; margin:2px 0; }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="main-header">
  <h2>🏭 Line Supply SOP — Direct Supply of Parts from Stores to Assy Line</h2>
  <p>Fill in all fields below, then generate your PDF & Excel report for download.</p>
</div>
""", unsafe_allow_html=True)

# ── SECTION 1: SOP Header ────────────────────────────────────────────────────
st.markdown('<div class="section-box"><div class="section-title">📄 Section 1: SOP Header Information</div>', unsafe_allow_html=True)
c1,c2,c3,c4 = st.columns(4)
with c1:
    company  = st.text_input("Company Name",  value="PINNACLE MOBILITY")
    sop_no   = st.text_input("SOP No.",        value="SCM/STR/LS/02")
with c2:
    brand    = st.text_input("Brand / Logo",   value="eka")
    rev_no   = st.text_input("Rev No.",        value="0.0")
with c3:
    unit     = st.text_input("Unit",           value="Chakan Plant")
    area     = st.text_input("Area",           value="Stores")
with c4:
    sub_area = st.text_input("Sub Area",       value="Line Side")
    zone     = st.text_input("Zone",           value="DIRECT LOCATIONS")

c5,c6 = st.columns(2)
with c5:
    sop_date  = st.date_input("SOP Date",  value=datetime.date(2024,12,26))
    page_info = st.text_input("Page Info", value="1 OUT OF 1")
with c6:
    purpose   = st.text_input("Purpose",   value="Supply of Supplier Packs to Line Side")
    scope     = st.text_area("Scope",      value="All Parts under direct supply category ( suggested by quality/operation)", height=70)
st.markdown('</div>', unsafe_allow_html=True)

# ── SECTION 2: Process Owner ─────────────────────────────────────────────────
st.markdown('<div class="section-box"><div class="section-title">👤 Section 2: Process Owner</div>', unsafe_allow_html=True)
c1,c2,c3 = st.columns(3)
with c1: owner           = st.text_input("Process Owner",        value="Stores Manager")
with c2: responsible     = st.text_input("Responsible Person",   value="Line Supervisor")
with c3: target_coverage = st.number_input("MTRL Coverage Target (%)", min_value=1, max_value=100, value=100)
st.markdown('</div>', unsafe_allow_html=True)

# ── SECTION 3: Parts ─────────────────────────────────────────────────────────
st.markdown('<div class="section-box"><div class="section-title">🔩 Section 3: Part Details & Coverage</div>', unsafe_allow_html=True)
st.caption("Add each part with its availability status. This drives the process flow decisions.")

if "parts" not in st.session_state:
    st.session_state.parts = [{"part_no":"","part_name":"","supplier":"","n_available":"Yes",
                                "n1_available":"Yes","shortfall":"","location":"","trolley_available":"Yes"}]

def add_part():    st.session_state.parts.append({"part_no":"","part_name":"","supplier":"","n_available":"Yes","n1_available":"Yes","shortfall":"","location":"","trolley_available":"Yes"})
def remove_part(i): st.session_state.parts.pop(i)

for i, part in enumerate(st.session_state.parts):
    with st.expander(f"Part {i+1}: {part['part_name'] or 'New Part'}", expanded=(i==0)):
        r1,r2,r3 = st.columns(3)
        with r1: st.session_state.parts[i]["part_no"]            = st.text_input("Part No.",          value=part["part_no"],            key=f"pno_{i}")
        with r2: st.session_state.parts[i]["part_name"]          = st.text_input("Part Name",         value=part["part_name"],          key=f"pnm_{i}")
        with r3: st.session_state.parts[i]["supplier"]           = st.text_input("Supplier",          value=part["supplier"],           key=f"sup_{i}")
        a1,a2,a3,a4 = st.columns(4)
        with a1: st.session_state.parts[i]["n_available"]        = st.selectbox("Available N Day?",   ["Yes","No"], index=0 if part["n_available"]=="Yes" else 1, key=f"nav_{i}")
        with a2: st.session_state.parts[i]["n1_available"]       = st.selectbox("Available N+1 Day?", ["Yes","No"], index=0 if part["n1_available"]=="Yes" else 1, key=f"n1v_{i}")
        with a3: st.session_state.parts[i]["trolley_available"]  = st.selectbox("In Loaded Trolley?", ["Yes","No"], index=0 if part["trolley_available"]=="Yes" else 1, key=f"tav_{i}")
        with a4: st.session_state.parts[i]["shortfall"]          = st.text_input("Shortfall Qty",     value=part["shortfall"],          key=f"sfl_{i}")
        st.session_state.parts[i]["location"] = st.text_input("Direct Part Location", value=part["location"], key=f"loc_{i}")
        if st.button(f"🗑 Remove Part {i+1}", key=f"rm_{i}"):
            remove_part(i); st.rerun()

st.button("➕ Add Another Part", on_click=add_part)
st.markdown('</div>', unsafe_allow_html=True)

# ── SECTION 4: Change Record ──────────────────────────────────────────────────
st.markdown('<div class="section-box"><div class="section-title">📝 Section 4: SOP Change Record</div>', unsafe_allow_html=True)

if "changes" not in st.session_state:
    st.session_state.changes = [{"eff_date":datetime.date(2024,12,17),"rev":"0.0","desc":"Original Version",
                                  "change_letter":"NA","prepared":"Prince S","reviewed":"Ajay G","approved":"Vilas B"}]

def add_change(): st.session_state.changes.append({"eff_date":datetime.date.today(),"rev":"","desc":"","change_letter":"","prepared":"","reviewed":"","approved":""})
def rm_change(i): st.session_state.changes.pop(i)

for i, ch in enumerate(st.session_state.changes):
    cc1,cc2,cc3,cc4,cc5,cc6,cc7,cc8 = st.columns([1.3,0.8,2,1,1,1,1,0.5])
    with cc1: st.session_state.changes[i]["eff_date"]      = st.date_input("Effective Date", value=ch["eff_date"],   key=f"cd_{i}")
    with cc2: st.session_state.changes[i]["rev"]           = st.text_input("Rev",            value=ch["rev"],        key=f"cr_{i}")
    with cc3: st.session_state.changes[i]["desc"]          = st.text_input("Description",    value=ch["desc"],       key=f"cds_{i}")
    with cc4: st.session_state.changes[i]["prepared"]      = st.text_input("Prepared By",    value=ch["prepared"],   key=f"cp_{i}")
    with cc5: st.session_state.changes[i]["reviewed"]      = st.text_input("Reviewed By",    value=ch["reviewed"],   key=f"cvw_{i}")
    with cc6: st.session_state.changes[i]["approved"]      = st.text_input("Approved By",    value=ch["approved"],   key=f"ca_{i}")
    with cc7: st.session_state.changes[i]["change_letter"] = st.text_input("Change Letter",  value=ch.get("change_letter","NA"), key=f"cl_{i}")
    with cc8:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🗑", key=f"rmc_{i}"): rm_change(i); st.rerun()

st.button("➕ Add Change Record", on_click=add_change)
st.markdown('</div>', unsafe_allow_html=True)

# ── Process Flow Preview ──────────────────────────────────────────────────────
with st.expander("🔍 Process Flow Preview", expanded=False):
    st.markdown(f'<div class="flow-step">🔵 <b>START</b> — Production lay sequence for N+1 day</div>', unsafe_allow_html=True)
    st.markdown('<div class="flow-arrow">↓</div>', unsafe_allow_html=True)
    st.markdown(f'<div class="flow-step">📋 <b>Step 1</b> — Verify line side stock & Update Direct Coverage Sheet <span style="color:#888;font-size:11px;">({responsible})</span></div>', unsafe_allow_html=True)
    st.markdown('<div class="flow-arrow">↓</div>', unsafe_allow_html=True)
    st.markdown(f'<div class="flow-step">📋 <b>Step 2</b> — Check material coverage w.r.t. N (today) day plan <span style="color:#888;font-size:11px;">({responsible})</span></div>', unsafe_allow_html=True)
    st.markdown('<div class="flow-arrow">↓</div>', unsafe_allow_html=True)
    st.markdown('<div class="flow-decision">🔷 <b>Decision 1</b> — Is Part available for N day?<br><span style="color:green;">YES →</span> Check N+1 coverage &nbsp;|&nbsp; <span style="color:red;">NO →</span> Check Loaded Trolley Area</div>', unsafe_allow_html=True)
    st.markdown('<div class="flow-arrow">↓ (YES)</div>', unsafe_allow_html=True)
    st.markdown(f'<div class="flow-step">📋 Check material coverage w.r.t. N+1 (tomorrow) day plan <span style="color:#888;font-size:11px;">({responsible})</span></div>', unsafe_allow_html=True)
    st.markdown('<div class="flow-arrow">↓</div>', unsafe_allow_html=True)
    st.markdown('<div class="flow-decision">🔷 <b>Decision 2</b> — Is Part available for N+1 day?<br><span style="color:green;">YES →</span> Continue &nbsp;|&nbsp; <span style="color:red;">NO →</span> Check Loaded Trolley</div>', unsafe_allow_html=True)
    st.markdown('<div class="flow-arrow">↓ (NO from D1 or D2)</div>', unsafe_allow_html=True)
    st.markdown(f'<div class="flow-step">📋 <b>Step 3</b> — Check part availability in Loaded Trolley Area <span style="color:#888;font-size:11px;">({responsible})</span></div>', unsafe_allow_html=True)
    st.markdown('<div class="flow-arrow">↓</div>', unsafe_allow_html=True)
    st.markdown('<div class="flow-decision">🔷 <b>Decision 3</b> — Is available in Loaded Trolley Area?<br><span style="color:green;">YES (A) →</span> Move Trolley & Update Bincard &nbsp;|&nbsp; <span style="color:red;">NO (B) →</span> Inform Responsible Person</div>', unsafe_allow_html=True)
    st.markdown('<div class="flow-arrow">↓</div>', unsafe_allow_html=True)
    st.markdown('<div class="flow-decision">🔷 <b>Decision 4</b> — Any more shortfall?<br><span style="color:green;">NO →</span> END &nbsp;|&nbsp; <span style="color:red;">YES →</span> Loop back to Step 3</div>', unsafe_allow_html=True)
    st.markdown('<div class="flow-arrow">↓ (NO)</div>', unsafe_allow_html=True)
    st.markdown('<div class="flow-step" style="background:#d4edda;border-left-color:#28a745;">🔴 <b>END</b> — Process Complete. No more shortfall.</div>', unsafe_allow_html=True)

# ════════════════════════════════════════════════════════════════
# PDF GENERATOR
# ════════════════════════════════════════════════════════════════
def generate_pdf():
    buf = io.BytesIO()
    W, H = A4
    c = rl_canvas.Canvas(buf, pagesize=A4)
    ml=18; pw=W-36

    def draw_diamond(cx, cy, hw, hh):
        p = c.beginPath()
        p.moveTo(cx,cy+hh); p.lineTo(cx+hw,cy); p.lineTo(cx,cy-hh); p.lineTo(cx-hw,cy); p.close()
        c.drawPath(p, stroke=1, fill=0)

    def arrow_dn(x, y, ln=9):
        c.line(x,y,x,y-ln)
        c.line(x-3,y-ln+4,x,y-ln); c.line(x+3,y-ln+4,x,y-ln)

    c.setLineWidth(1)
    # Header
    hh=42; ht=H-18
    c.rect(ml,ht-hh,pw,hh)
    lw=pw*0.28; c.rect(ml,ht-hh,lw,hh)
    c.setFont("Helvetica-Bold",8); c.drawString(ml+4,ht-12,company)
    c.setFillColorRGB(0,0.55,0.55); c.setFont("Helvetica-Bold",13)
    c.drawString(ml+4,ht-29,f"€ {brand}"); c.setFillColorRGB(0,0,0)
    tw2=pw*0.44; tx=ml+lw; c.rect(tx,ht-hh,tw2,hh)
    c.setFont("Helvetica-Bold",12)
    t="STANDARD OPERATING PROCEDURE"
    c.drawString(tx+(tw2-c.stringWidth(t,"Helvetica-Bold",12))/2,ht-14,t)
    c.setFont("Helvetica",7.5)
    s="Direct Supply of Parts from Stores to Assy Line"
    c.drawString(tx+(tw2-c.stringWidth(s,"Helvetica",7.5))/2,ht-26,s)
    ix=tx+tw2; iw=pw-lw-tw2; rh=hh/4
    lbs=[("SOP No.",sop_no),("Rev No.",rev_no),("Unit",unit),("Sub Area",sub_area)]
    rvs=[("Page",page_info),("Date",sop_date.strftime("%d-%m-%Y")),("Area",area),("Zone",zone)]
    for i,((l1,v1),(l2,v2)) in enumerate(zip(lbs,rvs)):
        yr=ht-hh+(3-i)*rh
        c.rect(ix,yr,iw/2,rh); c.rect(ix+iw/2,yr,iw/2,rh)
        c.setFont("Helvetica-Bold",5); c.drawString(ix+2,yr+rh/2+1,l1)
        c.setFont("Helvetica",5);      c.drawString(ix+2,yr+1,v1)
        c.setFont("Helvetica-Bold",5); c.drawString(ix+iw/2+2,yr+rh/2+1,l2)
        c.setFont("Helvetica",5);      c.drawString(ix+iw/2+2,yr+1,v2)

    # Purpose/Scope
    pt=ht-hh; ph=22; c.rect(ml,pt-ph,pw,ph)
    pw1=42; c.rect(ml,pt-ph,pw1,ph)
    c.setFont("Helvetica-Bold",8); c.drawString(ml+3,pt-ph/2-3,"Purpose")
    pv=pw*0.38; c.rect(ml+pw1,pt-ph,pv,ph)
    c.setFont("Helvetica",7); c.drawString(ml+pw1+4,pt-ph/2-3,purpose[:65])
    slx=ml+pw1+pv; slw=30; c.rect(slx,pt-ph,slw,ph)
    c.setFont("Helvetica-Bold",8); c.drawString(slx+3,pt-ph/2-3,"Scope")
    svx=slx+slw; svw=pw-pw1-pv-slw; c.rect(svx,pt-ph,svw,ph)
    c.setFont("Helvetica",6)
    sc1=scope[:58]; sc2=scope[58:116]
    c.drawString(svx+3,pt-10,sc1)
    if sc2: c.drawString(svx+3,pt-18,sc2)

    # Process Steps header
    pst=pt-ph; psth=14; c.rect(ml,pst-psth,pw,psth)
    c.setFont("Helvetica-Bold",9); c.drawString(ml+4,pst-10,"Process Steps:")
    ox=ml+pw*0.62; c.drawString(ox+10,pst-10,"OWNER :")
    c.drawString(ox+55,pst-10,owner)

    # Column headers
    ct=pst-psth; chh=18
    ciw=pw*0.12; cfw=pw*0.50; cow=pw*0.10; crw=pw*0.10; cdw=pw*0.09; cmw=pw-ciw-cfw-cow-crw-cdw
    col_defs=[(ml,ciw,"Input"),(ml+ciw,cfw,"Process Flow"),(ml+ciw+cfw,cow,"Output"),
              (ml+ciw+cfw+cow,crw,"Responsible"),(ml+ciw+cfw+cow+crw,cdw,"Doc. Format /\nSystem"),
              (ml+ciw+cfw+cow+crw+cdw,cmw,"Effective\nMeasurement")]
    for (cx,cw,lbl) in col_defs:
        c.rect(cx,ct-chh,cw,chh); c.setFont("Helvetica-Bold",6.5)
        lines=lbl.split('\n')
        if len(lines)==1:
            c.drawString(cx+(cw-c.stringWidth(lbl,"Helvetica-Bold",6.5))/2,ct-chh/2-3,lbl)
        else:
            c.drawString(cx+(cw-c.stringWidth(lines[0],"Helvetica-Bold",6.5))/2,ct-7,lines[0])
            c.drawString(cx+(cw-c.stringWidth(lines[1],"Helvetica-Bold",6.5))/2,ct-14,lines[1])

    # Body
    bt=ct-chh; bh=300
    for (cx,cw,_) in col_defs: c.rect(cx,bt-bh,cw,bh)
    c.setFont("Helvetica",6)
    c.drawString(ml+2,bt-80,"Direct Master Sheet")
    c.drawString(ml+2,bt-110,"Production Plan")
    c.drawString(ml+ciw+cfw+3,bt-105,"Coverage Shortfall")
    rsx=ml+ciw+cfw+cow
    for dy in [73,103,145,185,228]:
        c.setFont("Helvetica",5.5); c.drawString(rsx+1,bt-dy,responsible)
    c.setFont("Helvetica",6); c.drawString(rsx+crw+2,bt-103,"Hardcopy sheet")
    c.drawString(rsx+crw+cdw+2,bt-55,"MTRL Coverage (N+1)")
    c.drawString(rsx+crw+cdw+2,bt-64,f"(TARGET: {target_coverage}%)")

    # Flowchart
    fc_cx=ml+ciw+cfw/2; c.setLineWidth(0.7)
    o1cy=bt-22; ohw=85; ohh=8
    c.ellipse(fc_cx-ohw,o1cy-ohh,fc_cx+ohw,o1cy+ohh)
    c.setFont("Helvetica",6)
    t="Production lay sequence for N+1 day"
    c.drawString(fc_cx-c.stringWidth(t,"Helvetica",6)/2,o1cy-3,t)
    arrow_dn(fc_cx,o1cy-ohh)

    r1y=o1cy-ohh-9-12; r1w=140; r1h=12
    c.rect(fc_cx-r1w/2,r1y,r1w,r1h); c.setFont("Helvetica",5.6)
    t="Verify line side stock & Update Direct Coverage Sheet"
    c.drawString(fc_cx-c.stringWidth(t,"Helvetica",5.6)/2,r1y+3,t)
    arrow_dn(fc_cx,r1y)

    r2y=r1y-9-12; r2w=140; r2h=12
    c.rect(fc_cx-r2w/2,r2y,r2w,r2h); c.setFont("Helvetica",5.6)
    t="Check material coverage w.r.t. N (today) day plan"
    c.drawString(fc_cx-c.stringWidth(t,"Helvetica",5.6)/2,r2y+3,t)
    arrow_dn(fc_cx,r2y)

    d1cy=r2y-9-16; d1hw=48; d1hh=16
    draw_diamond(fc_cx,d1cy,d1hw,d1hh)
    c.setFont("Helvetica",5.5)
    c.drawString(fc_cx-c.stringWidth("Is Part available for","Helvetica",5.5)/2,d1cy+3,"Is Part available for")
    c.drawString(fc_cx-c.stringWidth("N day ?","Helvetica",5.5)/2,d1cy-4,"N day ?")

    ybx=fc_cx+d1hw+5; ybw=105; ybh=20
    c.line(fc_cx+d1hw,d1cy,ybx,d1cy)
    c.setFont("Helvetica",5); c.drawString(fc_cx+d1hw+1,d1cy+1,"YES")
    c.rect(ybx,d1cy-ybh/2,ybw,ybh); c.setFont("Helvetica",5.5)
    c.drawString(ybx+3,d1cy+3,"Check material coverage w.r.t. N+1")
    c.drawString(ybx+3,d1cy-4,"(tomorrow) day plan")

    d2cx=ybx+ybw+32; d2cy=d1cy; d2hw=20; d2hh=12
    c.line(ybx+ybw,d1cy,d2cx-d2hw,d1cy)
    draw_diamond(d2cx,d2cy,d2hw,d2hh)
    c.setFont("Helvetica",4.8)
    c.drawString(d2cx-c.stringWidth("Is Part available","Helvetica",4.8)/2,d2cy+3,"Is Part available")
    c.drawString(d2cx-c.stringWidth("for N+1 day ?","Helvetica",4.8)/2,d2cy-3,"for N+1 day ?")

    r4y=d1cy-d1hh-6-12; r4w=140; r4h=12
    c.line(d2cx-d2hw,d2cy,fc_cx+r4w/2+8,d2cy)
    c.line(fc_cx+r4w/2+8,d2cy,fc_cx+r4w/2+8,r4y+r4h/2)
    c.line(fc_cx+r4w/2+8,r4y+r4h/2,fc_cx+r4w/2,r4y+r4h/2)
    c.setFont("Helvetica",5); c.drawString(d2cx-d2hw-6,d2cy+1,"No")
    arrow_dn(fc_cx,d1cy-d1hh)
    c.setFont("Helvetica",5); c.drawString(fc_cx+2,d1cy-d1hh-4,"No")

    c.rect(fc_cx-r4w/2,r4y,r4w,r4h); c.setFont("Helvetica",5.6)
    t="Check the part availability in Loaded Trolley Area"
    c.drawString(fc_cx-c.stringWidth(t,"Helvetica",5.6)/2,r4y+3,t)
    arrow_dn(fc_cx,r4y)

    d3cy=r4y-6-14; d3hw=46; d3hh=14
    draw_diamond(fc_cx,d3cy,d3hw,d3hh)
    c.setFont("Helvetica",5.5)
    c.drawString(fc_cx-c.stringWidth("Is available in","Helvetica",5.5)/2,d3cy+4,"Is available in")
    c.drawString(fc_cx-c.stringWidth("Loaded Trolley","Helvetica",5.5)/2,d3cy-2,"Loaded Trolley")
    c.drawString(fc_cx-c.stringWidth("Area ?","Helvetica",5.5)/2,d3cy-8,"Area ?")

    bbx=fc_cx+d3hw+10; bbw=82; bbh=22
    c.line(fc_cx+d3hw,d3cy,bbx,d3cy)
    c.setFont("Helvetica",5); c.drawString(fc_cx+d3hw+1,d3cy+1,"No")
    c.rect(bbx,d3cy-bbh/2,bbw,bbh); c.setFont("Helvetica",5.5)
    c.drawString(bbx+3,d3cy-bbh/2+14,"B. Inform Responsible person &")
    c.drawString(bbx+3,d3cy-bbh/2+7,"seek for readiness confirmation")

    arrow_dn(fc_cx,d3cy-d3hh)
    c.setFont("Helvetica",5); c.drawString(fc_cx+2,d3cy-d3hh-4,"yes")
    ay=d3cy-d3hh-6-22; aw=100; ah=22; ax=fc_cx-aw-2
    c.rect(ax,ay,aw,ah); c.setFont("Helvetica",5.3)
    c.drawString(ax+3,ay+14,"A. Move Loaded Trolley to Direct Part")
    c.drawString(ax+3,ay+7,"Location & Update Bincard & Coverage sheet")
    c.line(ax+aw,ay+ah/2,bbx,d3cy-bbh/2+bbh/2)

    lx=ax-8
    c.line(fc_cx-d3hw,d3cy,lx,d3cy); c.line(lx,d3cy,lx,ay+ah)
    c.line(lx,ay+ah,ax,ay+ah/2)
    c.line(lx,ay,lx,ay-8)
    c.setFont("Helvetica",5.5); c.drawString(lx-22,ay-4,"YES")

    dfc_cx=fc_cx-10; dfc_cy=ay-8-14; dfc_hw=55; dfc_hh=14
    c.line(lx,ay,lx,dfc_cy); c.line(lx,dfc_cy,dfc_cx-dfc_hw,dfc_cy)
    draw_diamond(dfc_cx,dfc_cy,dfc_hw,dfc_hh)
    c.setFont("Helvetica",5.5)
    t="Any more shortfall as per?"
    c.drawString(dfc_cx-c.stringWidth(t,"Helvetica",5.5)/2,dfc_cy+3,t)

    fnoy=dfc_cy-dfc_hh; c.line(dfc_cx,fnoy,dfc_cx,fnoy-8)
    c.setFont("Helvetica",5); c.drawString(dfc_cx+1,fnoy-5,"No")
    ecy=fnoy-8-8
    c.setFillColorRGB(0.1,0.1,0.1)
    c.ellipse(dfc_cx-28,ecy-7,dfc_cx+28,ecy+7,stroke=1,fill=1)
    c.setFillColorRGB(0,0,0)

    # Change Record
    crt=bt-bh-4; crth=12
    c.rect(ml,crt-crth,pw,crth); c.setFont("Helvetica-Bold",8)
    t="SOP Change Record"
    c.drawString(ml+(pw-c.stringWidth(t,"Helvetica-Bold",8))/2,crt-9,t)
    chg_cols=[(30,"S.No."),(50,"EFFECTIVE DATE"),(38,"REV. No."),(130,"CHANGE DESCRIPTION"),
              (100,"CHANGE LETTER\n(P/D/S)"),(65,"PREPARED BY"),(65,"REVIEWED BY"),
              (pw-30-50-38-130-100-65-65,"APPROVED BY")]
    rch=12; cx2=ml
    for (cw,lbl) in chg_cols:
        c.rect(cx2,crt-crth-rch,cw,rch); c.setFont("Helvetica-Bold",5)
        lines=lbl.split('\n')
        if len(lines)==1:
            c.drawString(cx2+(cw-c.stringWidth(lbl,"Helvetica-Bold",5))/2,crt-crth-rch/2-2,lbl)
        else:
            c.drawString(cx2+(cw-c.stringWidth(lines[0],"Helvetica-Bold",5))/2,crt-crth-6,lines[0])
            c.drawString(cx2+(cw-c.stringWidth(lines[1],"Helvetica-Bold",5))/2,crt-crth-11,lines[1])
        cx2+=cw

    for r_i, ch_rec in enumerate(st.session_state.changes):
        row_data=[str(r_i+1),ch_rec["eff_date"].strftime("%d-%m-%Y"),ch_rec["rev"],
                  ch_rec["desc"][:40],ch_rec.get("change_letter","NA"),
                  ch_rec["prepared"],ch_rec["reviewed"],ch_rec["approved"]]
        cx2=ml; ry=crt-crth-rch*(r_i+2)
        for j,(cw,_) in enumerate(chg_cols):
            c.rect(cx2,ry,cw,rch); c.setFont("Helvetica",5.5)
            tw=c.stringWidth(row_data[j],"Helvetica",5.5)
            c.drawString(cx2+(cw-tw)/2,ry+3,row_data[j]); cx2+=cw

    for r_i in range(len(st.session_state.changes), len(st.session_state.changes)+2):
        cx2=ml; ry=crt-crth-rch*(r_i+2)
        for (cw,_) in chg_cols: c.rect(cx2,ry,cw,rch); cx2+=cw

    fy=crt-crth-rch*(len(st.session_state.changes)+4)-6
    c.setFont("Helvetica",6)
    ft="Composed By: Agilomatrix Pvt Ltd (connectus@agilomatrix.com)"
    c.drawString(ml+(pw-c.stringWidth(ft,"Helvetica",6))/2,fy,ft)

    c.save(); buf.seek(0)
    return buf

# ════════════════════════════════════════════════════════════════
# EXCEL GENERATOR
# ════════════════════════════════════════════════════════════════
def generate_excel():
    wb = openpyxl.Workbook()
    thin = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))

    def hcell(ws, row, col, val, bg="003366", fg="FFFFFF", bold=True, size=10, merge_end=None):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(color=fg, bold=bold, size=size)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
        if merge_end:
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_end)
        return cell

    def dcell(ws, row, col, val, bg=None, fc="000000", bold=False):
        cell = ws.cell(row=row, column=col, value=val)
        if bg: cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(color=fc, bold=bold, size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
        return cell

    # ── Sheet 1: SOP Summary ─────────────────────────────────
    ws = wb.active; ws.title = "SOP Summary"
    hcell(ws,1,1,"STANDARD OPERATING PROCEDURE — Direct Supply of Parts from Stores to Assy Line",merge_end=4,size=13)
    ws.row_dimensions[1].height = 30; ws.column_dimensions["A"].width=22; ws.column_dimensions["B"].width=35
    ws.column_dimensions["C"].width=22; ws.column_dimensions["D"].width=22

    fields=[("Company",company),("Brand",brand),("SOP No.",sop_no),("Rev No.",rev_no),
            ("Date",sop_date.strftime("%d-%m-%Y")),("Unit",unit),("Area",area),
            ("Sub Area",sub_area),("Zone",zone),("Page",page_info),
            ("Purpose",purpose),("Scope",scope),("Owner",owner),
            ("Responsible",responsible),("MTRL Target",f"{target_coverage}%")]
    for i,(k,v) in enumerate(fields,2):
        c1=ws.cell(row=i,column=1,value=k); c1.font=Font(bold=True,size=9,color="003366")
        c1.fill=PatternFill("solid",fgColor="CCE0FF"); c1.border=thin
        c1.alignment=Alignment(vertical="center")
        ws.merge_cells(start_row=i,start_column=2,end_row=i,end_column=4)
        c2=ws.cell(row=i,column=2,value=v); c2.border=thin
        c2.alignment=Alignment(vertical="center",wrap_text=True); ws.row_dimensions[i].height=16

    # ── Sheet 2: Parts Coverage ──────────────────────────────
    ws2=wb.create_sheet("Parts Coverage")
    hcell(ws2,1,1,"Parts Coverage Report — Line Supply SOP",merge_end=9,size=12)
    ws2.row_dimensions[1].height=24
    hdrs=["#","Part No.","Part Name","Supplier","Available N Day","Available N+1 Day","In Loaded Trolley","Shortfall Qty","Direct Part Location"]
    for j,h in enumerate(hdrs,1): hcell(ws2,2,j,h,bg="0055AA",size=9)
    ws2.row_dimensions[2].height=16

    for i,pt in enumerate(st.session_state.parts,1):
        n_ok=pt.get("n_available","Yes"); n1_ok=pt.get("n1_available","Yes"); tr_ok=pt.get("trolley_available","Yes")
        row=[i,pt.get("part_no",""),pt.get("part_name",""),pt.get("supplier",""),n_ok,n1_ok,tr_ok,pt.get("shortfall",""),pt.get("location","")]
        bg_alt="F0F6FF" if i%2==0 else "FFFFFF"
        for j,v in enumerate(row,1):
            if j in [5,6,7]:
                bg="D4EDDA" if v=="Yes" else "F8D7DA"; fc="155724" if v=="Yes" else "721C24"
                dcell(ws2,i+2,j,v,bg=bg,fc=fc,bold=True)
            else:
                dcell(ws2,i+2,j,v,bg=bg_alt)
        ws2.row_dimensions[i+2].height=15

    for j,w in enumerate([5,14,22,18,15,15,15,13,24],1): ws2.column_dimensions[get_column_letter(j)].width=w

    # ── Sheet 3: Process Flow ─────────────────────────────────
    ws3=wb.create_sheet("Process Flow")
    hcell(ws3,1,1,"Process Flow — Direct Supply SOP",merge_end=4,size=12)
    ws3.row_dimensions[1].height=24
    for j,h in enumerate(["Step","Type","Description","Responsible"],1): hcell(ws3,2,j,h,bg="0055AA",size=9)
    flow_steps=[
        ("START","Trigger","Production lay sequence for N+1 day","-"),
        ("Step 1","Process","Verify line side stock & Update Direct Coverage Sheet",responsible),
        ("Step 2","Process","Check material coverage w.r.t. N (today) day plan",responsible),
        ("Decision 1","Decision","Is Part available for N day? YES→Check N+1 | NO→Check Trolley",responsible),
        ("Step 3A","Process","Check material coverage w.r.t. N+1 (tomorrow) day plan",responsible),
        ("Decision 2","Decision","Is Part available for N+1? YES→Continue | NO→Check Trolley",responsible),
        ("Step 4","Process","Check the part availability in Loaded Trolley Area",responsible),
        ("Decision 3","Decision","Is available in Loaded Trolley Area? YES→Path A | NO→Path B",responsible),
        ("Path A","Process","Move Loaded Trolley to Direct Part Location & Update Bincard & Direct coverage sheet",responsible),
        ("Path B","Process","Inform Responsible person & seek for readiness confirmation",responsible),
        ("Decision 4","Decision","Any more shortfall? YES→Loop back | NO→END",responsible),
        ("END","End","Process Complete — No more shortfall","-"),
    ]
    type_colors={"Trigger":"E8F4FD","Process":"F0F6FF","Decision":"FDF5FF","End":"D4EDDA"}
    for i,(step,typ,desc,resp) in enumerate(flow_steps,3):
        bg=type_colors.get(typ,"FFFFFF")
        dcell(ws3,i,1,step,bg=bg,bold=True); dcell(ws3,i,2,typ,bg=bg)
        dcell(ws3,i,3,desc,bg=bg); dcell(ws3,i,4,resp,bg=bg)
        ws3.row_dimensions[i].height=22
    for j,w in enumerate([14,12,55,18],1): ws3.column_dimensions[get_column_letter(j)].width=w

    # ── Sheet 4: Change Record ────────────────────────────────
    ws4=wb.create_sheet("Change Record")
    hcell(ws4,1,1,"SOP Change Record",merge_end=8,size=12)
    ws4.row_dimensions[1].height=22
    ch_hdrs=["S.No.","Effective Date","Rev. No.","Change Description","Change Letter","Prepared By","Reviewed By","Approved By"]
    for j,h in enumerate(ch_hdrs,1): hcell(ws4,2,j,h,bg="0055AA",size=9)
    for i,ch in enumerate(st.session_state.changes,1):
        row=[i,ch["eff_date"].strftime("%d-%m-%Y"),ch["rev"],ch["desc"],ch.get("change_letter","NA"),ch["prepared"],ch["reviewed"],ch["approved"]]
        for j,v in enumerate(row,1): dcell(ws4,i+2,j,v,bg="F5F8FC" if i%2==0 else "FFFFFF")
        ws4.row_dimensions[i+2].height=15
    for j,w in enumerate([6,15,10,38,15,15,15,15],1): ws4.column_dimensions[get_column_letter(j)].width=w

    out=io.BytesIO(); wb.save(out); out.seek(0)
    return out

# ════════════════════════════════════════════════════════════════
# DOWNLOAD BUTTONS
# ════════════════════════════════════════════════════════════════
st.markdown("---")
st.subheader("📥 Generate & Download Reports")
g1,g2,g3=st.columns(3)

with g1:
    if st.button("🖨️ Generate PDF", use_container_width=True):
        with st.spinner("Building PDF..."):
            pdf_buf=generate_pdf()
        st.success("✅ PDF Ready!")
        st.download_button("⬇️ Download PDF",data=pdf_buf,
            file_name=f"Line_Supply_SOP_{sop_date.strftime('%d%m%Y')}.pdf",
            mime="application/pdf",use_container_width=True)

with g2:
    if st.button("📊 Generate Excel", use_container_width=True):
        with st.spinner("Building Excel..."):
            xl_buf=generate_excel()
        st.success("✅ Excel Ready!")
        st.download_button("⬇️ Download Excel",data=xl_buf,
            file_name=f"Line_Supply_SOP_{sop_date.strftime('%d%m%Y')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",use_container_width=True)

with g3:
    if st.button("📦 Generate Both", use_container_width=True):
        with st.spinner("Building reports..."):
            pdf_buf2=generate_pdf(); xl_buf2=generate_excel()
        st.success("✅ Both Ready!")
        st.download_button("⬇️ Download PDF",data=pdf_buf2,
            file_name=f"Line_Supply_SOP_{sop_date.strftime('%d%m%Y')}.pdf",
            mime="application/pdf",use_container_width=True)
        st.download_button("⬇️ Download Excel",data=xl_buf2,
            file_name=f"Line_Supply_SOP_{sop_date.strftime('%d%m%Y')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",use_container_width=True)

st.markdown("---")
st.markdown("<small style='color:#aaa;'>Composed By: Agilomatrix Pvt Ltd (connectus@agilomatrix.com) | Pinnacle Mobility — eka</small>", unsafe_allow_html=True)
